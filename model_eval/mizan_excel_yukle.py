#!/usr/bin/env python3
"""Sirkete ozel mizan.xlsx dosyasini PostgreSQL'e (mizan_alt_kirilim tablosu,
tenant_<vkn> semasi) yukler (2026-08-05, Excel'den DB'ye tasima).

Kullanici karari: "her kullanicinin farkli mizan bilgisi olduğu icin buna
gore bir yapi tasarlamamiz lazim" - core/mizan.py::get_alt_kirilimlar() artik
Excel degil bu tabloyu okuyor (bkz. core/mizan.py docstring'i).

Bir kereye mahsus/idempotent calistirilir: hedef semadaki mizan_alt_kirilim
tablosunu TRUNCATE edip Excel'i yeniden yukler - Excel guncellenirse script
tekrar calistirilarak DB senkron tutulur. Ayirici tutarsizligi (tire/bosluk)
core/mizan.py::_kod_normalize ile giderilir, duplike/bos kodlar loglanarak
atlanir (eski Excel-okuma davranisiyla AYNI, sadece hedef Excel'den DB'ye
degisti).

Kullanim (model_eval kokunden):

    DATABASE_URL=postgresql://efatura:sifre@localhost:5434/efatura_kdv \\
        python3 mizan_excel_yukle.py --vkn 0460351893 --excel exceller/mizan.xlsx
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from core.mizan import _kod_normalize

_logger = logging.getLogger(__name__)


def _mizan_satirlarini_oku(excel_yolu):
    """Excel'i satir satir okuyup (kod, ad) ciftlerine cevirir - HESAP KODU/
    HESAP ADI sutunlarini kullanir (bkz. Archive2/mizan.xlsx basligi, satir 6:
    'HESAP KODU', 'HESAP ADI', ...). Bu, core/mizan.py'nin eski (Excel-okuma
    donemindeki) _mizan_satirlarini_oku()'sunun BİREBİR aynısı - o fonksiyon
    DB'ye gecince kaldırıldı, mantık burada (tek seferlik yukleme aracı) yaşıyor."""
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_yolu), data_only=True)
    ws = wb[wb.sheetnames[0]]
    gorulen = set()
    satirlar = []
    for row_no, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        kod = row[0] if row else None
        ad = row[1] if row and len(row) > 1 else None
        if not kod or not isinstance(kod, str):
            continue
        kod_norm = _kod_normalize(kod)
        if not kod_norm:
            _logger.warning("%s satir %d bos/gecersiz kod, atlandi: %r", excel_yolu, row_no, kod)
            continue
        if kod_norm in gorulen:
            _logger.warning("%s satir %d duplike kod, atlandi: %s", excel_yolu, row_no, kod_norm)
            continue
        gorulen.add(kod_norm)
        satirlar.append((kod_norm, (ad or "").strip()))
    return satirlar


def yukle(vkn: str, excel_yolu: str, database_url: str | None = None) -> int:
    """Excel'deki tum satirlari tenant_<vkn>.mizan_alt_kirilim tablosuna
    yazar (TRUNCATE + INSERT). Sadece 3-seviyeli kodlar (XXX.YY.ZZZZZ) DB'ye
    yazilir - ust seviye ozet satirlari (sadece "191" gibi) atlanir, get_alt_kirilimlar
    onlari zaten kullanmiyordu (bkz. core/mizan.py eski gruplama mantigi)."""
    import psycopg2
    from psycopg2 import sql

    database_url = database_url or os.environ.get("DATABASE_URL")
    if not database_url:
        raise SystemExit(
            "DATABASE_URL env var tanimli degil - orn. "
            "postgresql://efatura:sifre@localhost:5434/efatura_kdv"
        )
    if not vkn.isdigit() or len(vkn) != 10:
        raise SystemExit(f"Gecersiz VKN: {vkn!r} - 10 haneli sayisal olmali")

    sema = f"tenant_{vkn}"
    satirlar = [
        (kod, kod.split(".")[0], ad)
        for kod, ad in _mizan_satirlarini_oku(excel_yolu)
        if len(kod.split(".")) == 3 and len(kod.split(".")[0]) == 3
    ]

    conn = psycopg2.connect(database_url)
    try:
        with conn.cursor() as cur:
            cur.execute(sql.SQL("SET search_path TO {}, public").format(sql.Identifier(sema)))
            cur.execute("TRUNCATE TABLE mizan_alt_kirilim")
            for hesap_kodu, ana_kod, hesap_adi in satirlar:
                cur.execute(
                    """
                    INSERT INTO mizan_alt_kirilim (hesap_kodu, ana_kod, hesap_adi)
                    VALUES (%s, %s, %s)
                    """,
                    (hesap_kodu, ana_kod, hesap_adi),
                )
        conn.commit()
    finally:
        conn.close()

    print(f"Tamamlandı: {sema}.mizan_alt_kirilim'e {len(satirlar)} satır yazıldı.")
    return len(satirlar)


def main() -> None:
    logging.basicConfig(level=logging.INFO)
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--vkn", required=True, help="Şirketin VKN'si (10 hane, tenant_<vkn> şeması hedeflenir)")
    ap.add_argument("--excel", required=True, help="mizan.xlsx dosyasının yolu")
    args = ap.parse_args()
    yukle(args.vkn, args.excel)


if __name__ == "__main__":
    main()
