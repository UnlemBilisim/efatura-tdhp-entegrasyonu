"""Mizan (hesap plani) alt kirilim referansi - TDHP_GLOSSARY'nin (genel,
tum muhasebeciler icin ayni 3 haneli kod aciklamalari) alt kirilim versiyonu.

Muhasebecinin kendi mizan.xlsx dosyasindan (Excel) her 3 haneli TDHP kodunun
o SIRKETE OZEL alt kirilimlarini (orn. 191.05.00005 = "%20 5/10 Tevkifatli
KDV") cikarir. TDHP_GLOSSARY'nin aksine bu liste sirketten sirkete FARKLIDIR
- 2026-07-24, kullanici karari: "alt kirilimlar her muhasebeci icin farkli
olabilir, sadece ana basliklar (101, 102 gibi) sabit".

Kullanim: core/single.py::predict_single_invoice() ana modelin urettigi
3 haneli kodu (orn. "191") aldiktan SONRA, bu kodun TUM alt kirilimlarini
(orn. 191'in altindaki 9 alt kod) ikinci bir LLM cagrisina (core/prompting.py::
build_alt_kirilim_prompt) verir - boylece alt kirilim secimi ayri, odakli
bir adimda yapilir, ana tahmin akisi (3 haneli kod bulma) etkilenmez."""

import threading
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_MIZAN_PATH = SCRIPT_DIR / "exceller" / "mizan.xlsx"

_mizan_cache = {}
_mizan_cache_lock = threading.Lock()


def _mizan_satirlarini_oku(mizan_path):
    """Excel'i satir satir okuyup (kod, ad) ciftlerine cevirir - HESAP KODU/
    HESAP ADI sutunlarini kullanir (bkz. Archive2/mizan.xlsx basligi,
    satir 6: 'HESAP KODU', 'HESAP ADI', ...)."""
    import openpyxl

    wb = openpyxl.load_workbook(str(mizan_path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    satirlar = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        kod = row[0] if row else None
        ad = row[1] if len(row) > 1 else None
        if not kod or not isinstance(kod, str):
            continue
        satirlar.append((kod.strip(), (ad or "").strip()))
    return satirlar


def get_alt_kirilimlar(mizan_path=None):
    """3 haneli TDHP koduna gore gruplanmis alt kirilim sozlugu doner:
    {"191": [("191.01.00020", "%20 Indirilecek KDV"), ...], ...}

    Sadece 3-seviyeli kodlar (XXX.YY.ZZZZZ formatinda, ana kodun DOGRUDAN
    alt kirilimlari) alinir - ust seviye ozet satirlari (orn. sadece "191"
    ya da "191.01") burada YOK, cunku onlar zaten TDHP_GLOSSARY'de ana
    kod olarak var; bu sozluk sadece "hangi alt kirilimdan birini secmen
    gerekiyor" sorusuna cevap veriyor.

    Process-omru cache'lenir (rag_common.py::get_collection ile ayni
    desen) - Excel her istekte yeniden okunmaz."""
    mizan_path = mizan_path or DEFAULT_MIZAN_PATH
    key = str(mizan_path)
    if key in _mizan_cache:
        return _mizan_cache[key]
    with _mizan_cache_lock:
        if key not in _mizan_cache:
            by_main_code = {}
            for kod, ad in _mizan_satirlarini_oku(mizan_path):
                parcalar = kod.split(".")
                if len(parcalar) == 3 and len(parcalar[0]) == 3:
                    ana_kod = parcalar[0]
                    by_main_code.setdefault(ana_kod, []).append((kod, ad))
            _mizan_cache[key] = by_main_code
    return _mizan_cache[key]


def reset_mizan_cache_for_tests():
    global _mizan_cache
    with _mizan_cache_lock:
        _mizan_cache = {}
